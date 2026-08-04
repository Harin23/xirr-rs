"""
Build a deterministic XIRR test corpus and an .xlsx workbook that computes
XIRR() for every case using the spreadsheet's own engine.

The workbook is recalculated by LibreOffice Calc (whose XIRR is the
Apache OpenOffice / LibreOffice `AnalysisAddIn::getXirr` implementation,
built for Excel compatibility). The resulting values become the golden file.

The same workbook can be opened in Excel or uploaded to Google Sheets to
regenerate the golden values from those engines.
"""
import random
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SEED = 20260803
random.seed(SEED)

ARIAL = "Arial"
HDR = Font(name=ARIAL, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F3864")
BODY = Font(name=ARIAL)
INPUT_FONT = Font(name=ARIAL, color="0000FF")   # blue = hardcoded input
NOTE = Font(name=ARIAL, italic=True, size=9)

cases = []  # (name, [(date, amount), ...])


def add(name, flows):
    cases.append((name, flows))


D0 = date(2015, 1, 1)


def d(days):
    return D0 + timedelta(days=days)


# --- 1. conventional, single root -------------------------------------------
add("conventional/textbook-3flow", [(date(2020, 1, 1), -1000), (date(2021, 1, 1), 750), (date(2022, 1, 1), 500)])
add("conventional/one-year-30pct", [(date(2020, 1, 1), -100), (date(2021, 1, 1), 130)])
add("conventional/monthly-12", [(date(2021, m, 1), -100 if m == 1 else 20) for m in range(1, 13)])
add("conventional/quarterly-drawdown", [
    (date(2019, 1, 1), -5000), (date(2019, 4, 1), -3000), (date(2019, 7, 1), -2000),
    (date(2020, 1, 1), 1500), (date(2021, 1, 1), 4000), (date(2022, 1, 1), 6500)])
add("conventional/leap-year-span", [(date(2019, 12, 31), -1000), (date(2020, 12, 31), 1150)])

# --- 2. horizon extremes ----------------------------------------------------
add("horizon/30y-50x", [(date(1990, 1, 1), -1000), (date(2020, 1, 1), 50000)])
add("horizon/50y-50x", [(date(1975, 1, 1), -1000), (date(2025, 1, 1), 50000)])
add("horizon/5-day-10pct", [(date(2020, 1, 1), -1000), (date(2020, 1, 6), 1100)])
add("horizon/1-day", [(date(2020, 1, 1), -1000), (date(2020, 1, 2), 1001)])
add("horizon/tiny-basis-10y", [(date(2020, 1, 1), -1), (date(2020, 1, 2), -1), (date(2030, 1, 1), 1000000)])

# --- 3. rate extremes -------------------------------------------------------
add("rate/near-zero-positive", [(date(2020, 1, 1), -1000), (date(2021, 1, 1), 1000.0001)])
add("rate/exact-zero", [(date(2020, 1, 1), -1000), (date(2021, 1, 1), 1000)])
add("rate/near-zero-negative", [(date(2020, 1, 1), -1000), (date(2021, 1, 1), 999.9999)])
add("rate/near-total-loss", [(date(2020, 1, 1), -1000), (date(2021, 1, 1), 1)])
add("rate/deep-negative", [(date(2020, 1, 1), -10000), (date(2021, 1, 1), 500), (date(2022, 1, 1), 250)])

# --- 4. multiple roots (the whole point) ------------------------------------
add("multiroot/classic-100-230-132", [(date(2020, 1, 1), -100), (date(2021, 1, 1), 230), (date(2022, 1, 1), -132)])
add("multiroot/sign-reversal-monthly", [(date(2020, 1, 1), -100), (date(2020, 2, 1), 10000), (date(2020, 3, 1), -9000)])
add("multiroot/three-roots", [(date(2015, 1, 1), -1000), (date(2016, 1, 1), 3000), (date(2017, 1, 1), -2500), (date(2018, 1, 1), 600)])
add("multiroot/pe-recall", [(date(2018, 1, 1), -1000), (date(2019, 1, 1), 5000), (date(2020, 1, 1), -6000), (date(2021, 1, 1), 2500)])
add("multiroot/five-flips", [(date(2020, 1, 1), -500), (date(2020, 7, 1), 3000), (date(2021, 1, 1), -5000),
                             (date(2021, 7, 1), 3000), (date(2022, 1, 1), -400)])
add("multiroot/capital-call-mid", [(date(2017, 3, 15), -2500), (date(2018, 6, 30), 4200),
                                   (date(2019, 2, 1), -3100), (date(2021, 9, 30), 2400)])

# --- 5. magnitude / scale ---------------------------------------------------
for exp in (0, 3, 6, 9, 12):
    k = 10 ** exp
    add(f"scale/1e{exp}-conventional", [(date(2020, 1, 1), -1000 * k), (date(2021, 1, 1), 750 * k), (date(2022, 1, 1), 500 * k)])
    add(f"scale/1e{exp}-multiroot", [(date(2020, 1, 1), -1000 * k), (date(2021, 1, 1), 5000 * k),
                                     (date(2022, 1, 1), -6000 * k), (date(2023, 1, 1), 2500 * k)])
add("scale/sub-cent", [(date(2020, 1, 1), -0.01), (date(2021, 1, 1), 0.0125)])

# --- 6. structural edge cases ----------------------------------------------
add("edge/zero-amounts-interleaved", [(date(2020, 1, 1), -1000), (date(2020, 6, 1), 0),
                                      (date(2021, 1, 1), 0), (date(2022, 1, 1), 1300)])
add("edge/duplicate-dates", [(date(2020, 1, 1), -500), (date(2020, 1, 1), -500), (date(2022, 1, 1), 1300)])
add("edge/many-flows-60", [(date(2020 + i // 12, i % 12 + 1, 1), -10000 if i == 0 else 175) for i in range(60)])
add("edge/trailing-zero", [(date(2020, 1, 1), -1000), (date(2021, 1, 1), 1200), (date(2022, 1, 1), 0)])

# --- 7. no root exists (expect #NUM!) ---------------------------------------
add("nosolution/lease-tail", (
    [(date(2026, 6, 4), -176), (date(2026, 6, 4), 25000)]
    + [(date(2026 + (m - 7) // 12 + (1 if (m - 7) % 12 >= 0 else 0), 0, 1), 0) for m in []]
    + [(date(2036, 6, 3), 433)]))
add("nosolution/all-tiny-outflow", [(date(2020, 1, 1), -0.000001), (date(2021, 1, 1), 100000), (date(2030, 1, 1), 5)])

# --- 8. randomised fuzz corpus ---------------------------------------------
for i in range(60):
    k = random.randint(3, 8)
    flows, cur = [], date(2015, 1, 1)
    for _ in range(k):
        cur = cur + timedelta(days=random.randint(30, 500))
        flows.append((cur, float(random.randint(-5000, 5000))))
    if not (any(a > 0 for _, a in flows) and any(a < 0 for _, a in flows)):
        continue
    add(f"fuzz/{i:03d}", flows)

# ---------------------------------------------------------------------------
wb = Workbook()

# ---- Sheet: README --------------------------------------------------------
ws = wb.active
ws.title = "README"
lines = [
    ("xirr-rs golden-value corpus", True),
    ("", False),
    ("HOW TO REGENERATE GOLDEN VALUES FROM A REAL SPREADSHEET ENGINE", True),
    ("1. Open this workbook in Excel, Google Sheets, or LibreOffice Calc.", False),
    ("2. Let it recalculate. The 'results' sheet column D computes XIRR() natively.", False),
    ("3. Copy 'results' columns A and D into golden_<engine>.csv (case_id,expected).", False),
    ("4. A cell showing the numeric-error marker means the engine could not solve it; record NUM.", False),
    ("", False),
    ("Column D formula:  =IFERROR(XIRR(amount_range, date_range), \"NUM\")", False),
    ("Guess is left at the engine default of 10%, matching production callers.", False),
    ("", False),
    ("Cash flows live on the 'flows' sheet in long format, one row per payment,", False),
    ("grouped by case_id and already in the exact input order the library receives.", False),
    ("Input order is significant: XIRR uses the FIRST row of each block as the", False),
    ("reference date, not the earliest date.", False),
    ("", False),
    (f"Deterministic seed: {SEED}", False),
    (f"Cases: {len(cases)}", False),
]
for r, (txt, bold) in enumerate(lines, start=1):
    c = ws.cell(row=r, column=1, value=txt)
    c.font = Font(name=ARIAL, bold=bold)
ws.column_dimensions["A"].width = 95

# ---- Sheet: flows ---------------------------------------------------------
fw = wb.create_sheet("flows")
for i, h in enumerate(["case_id", "date", "amount"], start=1):
    c = fw.cell(row=1, column=i, value=h)
    c.font, c.fill = HDR, HDR_FILL
row = 2
ranges = {}
for name, flows in cases:
    start = row
    for dt, amt in flows:
        fw.cell(row=row, column=1, value=name).font = BODY
        dc = fw.cell(row=row, column=2, value=dt)
        dc.font, dc.number_format = INPUT_FONT, "yyyy-mm-dd"
        ac = fw.cell(row=row, column=3, value=float(amt))
        ac.font, ac.number_format = INPUT_FONT, "General"
        row += 1
    ranges[name] = (start, row - 1)
fw.column_dimensions["A"].width = 32
fw.column_dimensions["B"].width = 13
fw.column_dimensions["C"].width = 18
fw.freeze_panes = "A2"

# ---- Sheet: results -------------------------------------------------------
rw = wb.create_sheet("results")
for i, h in enumerate(["case_id", "n_flows", "sign_changes", "xirr", "xnpv_at_xirr"], start=1):
    c = rw.cell(row=1, column=i, value=h)
    c.font, c.fill = HDR, HDR_FILL
for r, (name, flows) in enumerate(cases, start=2):
    a, b = ranges[name]
    amt = f"flows!$C${a}:$C${b}"
    dts = f"flows!$B${a}:$B${b}"
    sc = sum(1 for x, y in zip(flows, flows[1:])
             if x[1] != 0 and y[1] != 0 and (x[1] > 0) != (y[1] > 0))
    rw.cell(row=r, column=1, value=name).font = BODY
    rw.cell(row=r, column=2, value=len(flows)).font = BODY
    rw.cell(row=r, column=3, value=sc).font = BODY
    c = rw.cell(row=r, column=4, value=f'=IFERROR(XIRR({amt},{dts}),"NUM")')
    c.font, c.number_format = BODY, "0.000000000000"
    # residual check: XNPV evaluated at the rate the engine returned
    c2 = rw.cell(row=r, column=5, value=f'=IF(ISNUMBER(D{r}),IFERROR(XNPV(D{r},{amt},{dts}),"ERR"),"")')
    c2.font, c2.number_format = BODY, "0.00E+00"
rw.column_dimensions["A"].width = 32
rw.column_dimensions["B"].width = 10
rw.column_dimensions["C"].width = 14
rw.column_dimensions["D"].width = 22
rw.column_dimensions["E"].width = 16
rw.freeze_panes = "A2"
n = len(cases) + 2
rw.cell(row=n + 1, column=1, value="xirr uses the engine default guess (10%). NUM = the engine reported a numeric error (no solution found).").font = NOTE

wb.save("/home/claude/work/xirr_golden_corpus.xlsx")
print(f"cases={len(cases)} rows={row-1}")
