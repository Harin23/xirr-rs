"""Export cases.csv and expected_<engine>.csv from a recalculated workbook."""
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

WB = Path(sys.argv[1] if len(sys.argv) > 1 else "xirr_golden_corpus.xlsx")
ENGINE = sys.argv[2] if len(sys.argv) > 2 else "libreoffice"
OUT = Path("__test__/golden")
OUT.mkdir(parents=True, exist_ok=True)

wb = load_workbook(WB, data_only=True)
flows, order = {}, []
for cid, dt, amt in wb["flows"].iter_rows(min_row=2, values_only=True):
    if cid is None:
        continue
    if cid not in flows:
        flows[cid] = []
        order.append(cid)
    flows[cid].append((dt.date().isoformat(), repr(float(amt))))

with (OUT / "cases.csv").open("w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["case_id", "date", "amount"])
    for cid in order:
        for d, a in flows[cid]:
            w.writerow([cid, d, a])

gold = {r[0]: r[3] for r in wb["results"].iter_rows(min_row=2, values_only=True) if r[0] in flows}
missing = [c for c in order if c not in gold]
if missing:
    raise SystemExit(f"workbook not recalculated; {len(missing)} results are empty")

with (OUT / f"expected_{ENGINE}.csv").open("w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["case_id", "expected"])
    for cid in order:
        v = gold[cid]
        w.writerow([cid, "NUM" if v == "NUM" else repr(float(v))])

print(f"{len(order)} cases -> {OUT}/expected_{ENGINE}.csv")
